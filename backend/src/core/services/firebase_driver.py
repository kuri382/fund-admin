from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Literal, Optional

from fastapi import HTTPException, UploadFile
from google.cloud import firestore
from google.cloud.firestore_v1.base_query import FieldFilter
from openpyxl import load_workbook
from pydantic import BaseModel, Field, validator


async def upload_to_firebase(file: UploadFile, filename: str, storage_client):
    blob = storage_client.blob(filename)
    file_content = await file.read()
    blob.upload_from_string(
        file_content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    blob.make_public()
    return blob.public_url


async def parse_excel_file(file: UploadFile):
    file_content = await file.read()
    workbook = load_workbook(BytesIO(file_content), read_only=True)
    sheet = workbook.active
    data = [row for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return data


def get_project_id(user_id, firestore_client) -> str:
    projects_ref = firestore_client.collection('users').document(user_id).collection('projects')
    is_selected_filter = firestore.FieldFilter("is_selected", "==", True)
    query = projects_ref.where(filter=is_selected_filter).limit(1)
    selected_project = query.get()
    if not selected_project:
        raise ValueError("No project selected for the user")

    return str(selected_project[0].id)


class AnalysisResult(BaseModel):
    abstract: str
    feature: str
    extractable_info: list[str]
    year_info: str
    period_type: str
    category: str
    category_ir: str


class AnalystReport(BaseModel):
    facts: str =     Field(..., description='現状の確認事項や客観的なデータ')
    issues: str =    Field(..., description='潜在的なリスクや経営上の懸念点')
    rationale: str = Field(..., description='課題やリスクを推測した理由や思考プロセス')
    forecast: str =  Field(..., description='リスクが顕在化した場合の影響や将来の見通し')
    investigation: str =    Field(..., description='課題やリスク推測をより精緻に行うために必要な情報')


class TranscriptionReport(BaseModel):
    transcription: str =     Field(..., description='ファイルに記述されたすべての内容を正確かつ丁寧に抽出した文章')


def save_page_analyst_report(
    firestore_client: firestore.Client,
    user_id: str,
    file_uuid: str,
    file_name: str,
    page_number: int,
    analyst_report: AnalystReport,
    transcription_report: TranscriptionReport,
) -> None:
    projects_ref = firestore_client.collection('users').document(user_id).collection('projects')
    is_selected_filter = FieldFilter("is_selected", "==", True)
    query = projects_ref.where(filter=is_selected_filter).limit(1)
    selected_project = query.get()

    if not selected_project:
        raise ValueError("No project selected for the user")

    selected_project_id = selected_project[0].id
    doc_ref = (
        firestore_client.collection('users')
        .document(user_id)
        .collection('projects')
        .document(selected_project_id)
        .collection('documents')
        .document(str(file_uuid))
    )
    page_ref = doc_ref.collection('pages').document(str(page_number))

    try:
        page_ref.set(
            {
                # metadata
                "file_uuid": str(file_uuid),
                "file_name": file_name,
                "page_number": str(page_number),
                # report
                "facts": str(analyst_report.facts),
                "issues": str(analyst_report.issues),
                'rationale': str(analyst_report.rationale),
                'forecast': str(analyst_report.forecast),
                'investigation': str(analyst_report.investigation),
                'transcription': str(transcription_report.transcription)
            }
        )
        return

    except Exception as e:
        raise HTTPException(status_code=500, detail=e)


def save_analysis_result(
    firestore_client: firestore.Client,
    user_id: str,
    file_name: str,
    file_uuid: str,
    analysis_result: AnalysisResult,
    target_collection: str,
) -> None:
    projects_ref = firestore_client.collection('users').document(user_id).collection('projects')
    is_selected_filter = FieldFilter("is_selected", "==", True)
    query = projects_ref.where(filter=is_selected_filter).limit(1)
    selected_project = query.get()

    if not selected_project:
        raise ValueError("No project selected for the user")

    selected_project_id = selected_project[0].id
    doc_ref = (
        firestore_client.collection('users')
        .document(user_id)
        .collection('projects')
        .document(selected_project_id)
        .collection(target_collection)
        .document(str(file_uuid))
    )

    try:
        doc_ref.set(
            {
                "file_name": file_name,
                "file_uuid": str(file_uuid),
                "abstract": analysis_result.abstract,
                "feature": analysis_result.feature,
                "extractable_info": analysis_result.extractable_info,
                "year_info": analysis_result.year_info,
                "period_type": analysis_result.period_type,
                "category": analysis_result.category,
                "category_ir": analysis_result.category_ir,
            }
        )
        return

    except Exception as e:
        raise HTTPException(status_code=500, detail=e)


class Period(BaseModel):
    year: int = Field(..., description="年度を表す。例: 2024")
    month: int = Field(..., description="月を表す。例: 8")
    quarter: Optional[int] = Field(..., description="四半期を表す。例: 第2四半期なら2")
    period_type: Literal["年度", "月次", "四半期"] = Field(
        None, description="期間の種類を表す。'期' または '四半期' など"
    )

    @validator('year')
    def year_must_be_four_digits(cls, v):
        if not (1000 <= v <= 9999):
            raise ValueError("Year must be a four-digit integer")
        return v

    @validator('month')
    def month_must_be_valid(cls, v):
        if not (1 <= v <= 12):
            raise ValueError("Month must be between 1 and 12")
        return v

    @validator('quarter')
    def quarter_must_be_valid(cls, v):
        if v is not None and not (1 <= v <= 4):
            raise ValueError("Quarter must be between 1 and 4")
        return v


class BusinessSummary(BaseModel):
    period: Period

    # 売上高 Revenue
    revenue_forecast: Decimal | None = Field(..., description='売上高 予測。単位は円')
    revenue_actual: Decimal | None = Field(..., description='売上高 実績。単位は円')

    # 売上総利益 Gross Profit
    gross_profit_forecast: Decimal | None = Field(..., description='売上総利益 予測。単位は円')
    gross_profit_actual: Decimal | None = Field(..., description='売上総利益 実績。単位は円')

    # 売上総利益率
    gross_profit_margin_forecast: Decimal | None = Field(..., description='売上総利益率 予測。単位は円')
    gross_profit_margin_actual: Decimal | None = Field(..., description='売上総利益率 実績。単位は円')

    def has_non_none_fields(self) -> bool:
        """period以外のフィールドが全てNoneならFalse、少なくとも1つでもNoneでないフィールドがあればTrueを返す"""
        return any(value is not None for field, value in self.dict(exclude={'period'}).items())


def save_page_image_analysis(
    firestore_client: firestore.Client,
    user_id: str,
    file_uuid: str,
    file_name: str,
    page_uuid: str,
    page_number: int,
    business_summary: BusinessSummary,
    explanation: str,
    output: str,
    opinion: str,
    target_collection: str = 'sales',
) -> None:
    projects_ref = firestore_client.collection('users').document(user_id).collection('projects')
    is_selected_filter = FieldFilter("is_selected", "==", True)
    query = projects_ref.where(filter=is_selected_filter).limit(1)
    selected_project = query.get()

    if not selected_project:
        raise ValueError("No project selected for the user")

    selected_project_id = selected_project[0].id
    doc_ref = (
        firestore_client.collection('users')
        .document(user_id)
        .collection('projects')
        .document(selected_project_id)
        .collection(target_collection)
        .document(str(page_uuid))
    )

    try:
        doc_ref.set(
            {
                "file_uuid": str(file_uuid),
                "file_name": file_name,
                "page_number": page_number,
                "year": business_summary.period.year,
                "month": business_summary.period.month,
                "quarter": business_summary.period.quarter,
                "period_type": business_summary.period.period_type,
                "revenue_forecast": str(business_summary.revenue_forecast),
                "revenue_actual": str(business_summary.revenue_actual),
                "gross_profit_forecast": str(business_summary.gross_profit_forecast),
                "gross_profit_actual": str(business_summary.gross_profit_actual),
                "gross_profit_margin_forecast": str(business_summary.gross_profit_margin_forecast),
                "gross_profit_margin_actual": str(business_summary.gross_profit_margin_actual),
                "explanation": str(explanation),
                "output": str(output),
                'opinion': str(opinion),
            }
        )
        return

    except Exception as e:
        raise HTTPException(status_code=500, detail=e)


def safe_decimal(value: Optional[str]) -> Optional[Decimal]:
    """Convert a value to Decimal safely, return None if conversion fails."""
    if value is None or value == "":
        return None
    try:
        return Decimal(value)
    except InvalidOperation:
        return None


def get_selected_project_id(firestore_client: firestore.Client, user_id: str) -> str:
    projects_ref = firestore_client.collection('users').document(user_id).collection('projects')
    is_selected_filter = FieldFilter("is_selected", "==", True)
    query = projects_ref.where(filter=is_selected_filter).limit(1)
    selected_project = query.get()

    if not selected_project:
        raise ValueError("No project selected for the user")

    return selected_project[0].id


def parse_business_summary(doc) -> BusinessSummary:
    data = doc.to_dict()
    period = Period(
        year=data["year"],
        month=data["month"],
        quarter=data.get("quarter"),
        period_type=data.get("period_type"),
    )
    return BusinessSummary(
        period=period,
        revenue_forecast=safe_decimal(data.get("revenue_forecast")),
        revenue_actual=safe_decimal(data.get("revenue_actual")),
        gross_profit_forecast=safe_decimal(data.get("gross_profit_forecast")),
        gross_profit_actual=safe_decimal(data.get("gross_profit_actual")),
        gross_profit_margin_forecast=safe_decimal(data.get("gross_profit_margin_forecast")),
        gross_profit_margin_actual=safe_decimal(data.get("gross_profit_margin_actual")),
    )


def fetch_page_parameter_analysis(
    firestore_client: firestore.Client,
    user_id: str,
    page_uuid: Optional[str] = None,
    target_collection: str = 'sales',
) -> list[BusinessSummary]:
    selected_project_id = get_selected_project_id(firestore_client, user_id)
    collection_ref = (
        firestore_client.collection('users')
        .document(user_id)
        .collection('projects')
        .document(selected_project_id)
        .collection(target_collection)
    )

    if page_uuid:
        doc = collection_ref.document(page_uuid).get()
        return [parse_business_summary(doc)] if doc.exists else []
    else:
        return [parse_business_summary(doc) for doc in collection_ref.stream()]


class ResAnalystReportItem(BaseModel):
    page_number: int
    facts: str
    issues: str
    rationale: str
    forecast: str
    investigation: str
    transcription: str


def fetch_page_summary(
    firestore_client: firestore.Client,
    user_id: str,
    file_uuid: str,
    limit: int = 1000,
) -> list[ResAnalystReportItem]:
    # ユーザーの選択されたプロジェクトIDを取得
    selected_project_id = get_selected_project_id(firestore_client, user_id)

    # `file_uuid` に対応するドキュメントへの参照
    file_ref = (
        firestore_client.collection('users')
        .document(user_id)
        .collection('projects')
        .document(selected_project_id)
        .collection('documents')
        .document(file_uuid)
    )

    # `pages` サブコレクションを参照し、すべてのページデータを取得
    pages_ref = file_ref.collection('pages')
    query = pages_ref.limit(limit)
    page_docs = query.get()

    # 文字列型の `page_number` を昇順ソート
    sorted_docs = sorted(
        page_docs,
        key=lambda doc: int(doc.to_dict().get('page_number', 0))
    )

    # ページごとのデータをリストに変換
    parameter_summaries = []
    for doc in sorted_docs:
        page_data = doc.to_dict()
        parameter_summary = ResAnalystReportItem(
            page_number=page_data.get('page_number'),
            facts=page_data.get('facts'),
            issues=page_data.get('issues'),
            rationale=page_data.get('rationale'),
            forecast=page_data.get('forecast'),
            investigation=page_data.get('investigation'),
            transcription=page_data.get('transcription'),
        )
        parameter_summaries.append(parameter_summary)

    return parameter_summaries


def retrieve_and_convert_to_json(
    firestore_client, user_id: str, file_uuid: str, target_collection: str = 'plans'
) -> str:
    projects_ref = firestore_client.collection('users').document(user_id).collection('projects')
    is_selected_filter = FieldFilter("is_selected", "==", True)
    query = projects_ref.where(filter=is_selected_filter).limit(1)
    selected_project = query.get()

    if not selected_project:
        raise ValueError("No project selected for the user")
    selected_project_id = selected_project[0].id
    doc_ref = (
        firestore_client.collection('users')
        .document(user_id)
        .collection('projects')
        .document(selected_project_id)
        .collection(target_collection)
        .document(str(file_uuid))
    )

    try:
        doc = doc_ref.get()
        if not doc.exists:
            raise ValueError("Document does not exist")

        data = doc.to_dict()
        # json_data = json.dumps(data, default=str)  # Decimal型を文字列として出力
        return data

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class PageDetail(BaseModel):
    index: int
    summary: str
    updated_at: datetime


def save_pages_to_analysis_result(
    firestore_client: firestore.Client,
    user_id: str,
    file_uuid: str,
    target_collection: str,
    pages: list[PageDetail],
) -> None:
    projects_ref = firestore_client.collection('users').document(user_id).collection('projects')
    is_selected_filter = FieldFilter("is_selected", "==", True)
    query = projects_ref.where(filter=is_selected_filter).limit(1)
    selected_project = query.get()

    if not selected_project:
        raise ValueError("No project selected for the user")

    selected_project_id = selected_project[0].id
    doc_ref = (
        firestore_client.collection('users')
        .document(user_id)
        .collection('projects')
        .document(selected_project_id)
        .collection(target_collection)
        .document(str(file_uuid))
    )

    try:
        pages_dicts = [
            {
                "index": page.index,
                "summary": page.summary,
                "updated_at": page.updated_at,
            }
            for page in pages
        ]
        doc_ref.update({"pages": firestore.ArrayUnion(pages_dicts)})
        return

    except Exception as e:
        detail = 'saving pages detail to firebase'
        raise HTTPException(status_code=500, detail=f'{detail}: {e}')


async def get_pages_from_analysis_result(
    firestore_client: firestore.Client,
    user_id: str,
    file_uuid: str,
    target_collection: str,
) -> list[PageDetail]:
    """指定されたUUIDに紐づくデータを取得し、重複するindexがあれば最新のものを返す"""
    projects_ref = firestore_client.collection('users').document(user_id).collection('projects')
    is_selected_filter = firestore.FieldFilter("is_selected", "==", True)
    query = projects_ref.where(filter=is_selected_filter).limit(1)
    selected_project = query.get()

    if not selected_project:
        raise ValueError("No project selected for the user")

    selected_project_id = selected_project[0].id
    doc_ref = (
        firestore_client.collection('users')
        .document(user_id)
        .collection('projects')
        .document(selected_project_id)
        .collection(target_collection)
        .document(str(file_uuid))
    )

    # Firestoreからデータを取得
    doc_snapshot = doc_ref.get()
    if not doc_snapshot.exists:
        raise HTTPException(status_code=404, detail="No pages found for the specified UUID")

    # Firestoreから取得した`pages`フィールドのリストを取得
    pages_data = doc_snapshot.to_dict().get("pages", [])

    # `index`ごとに最新のページを取得
    latest_pages = {}
    for page in pages_data:
        index = page["index"]
        updated_at = datetime.fromtimestamp(page["updated_at"].timestamp())

        # `index` が存在しない、もしくは `updated_at` がより新しい場合に更新
        if index not in latest_pages or updated_at > latest_pages[index].updated_at:
            latest_pages[index] = PageDetail(index=index, summary=page["summary"], updated_at=updated_at)

    return list(latest_pages.values())


def save_worker_analyst_report(
    firestore_client: firestore.Client,
    user_id: str,
    project_id: str,
    file_uuid: str,
    result_sentence: str,
) -> None:
    doc_ref = (
        firestore_client.collection('users')
        .document(user_id)
        .collection('projects')
        .document(project_id)
        .collection('documents')
        .document(str(file_uuid))
    )

    # analyst_report コレクションの参照を取得
    analyst_report_ref = doc_ref.collection('analyst_report').document('summary')

    try:
        analyst_report_ref.set({
            "analyst_summary": result_sentence,
        })
        return

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
