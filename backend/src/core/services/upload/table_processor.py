import io
import csv

from openpyxl import load_workbook


def convert_xlsx_row_to_text(contents) -> str:
    file_stream = io.BytesIO(contents)
    workbook = load_workbook(file_stream)
    sheet = workbook.active

    sheet_content = []
    for row in sheet.iter_rows(values_only=True):
        sheet_content.append("\t".join([str(cell) for cell in row if cell is not None]))
    text_content = "\n".join(sheet_content)
    return text_content


def analyze_csv_content(contents: bytes) -> str:
    file_stream = io.StringIO(contents.decode("utf-8"))
    reader = csv.reader(file_stream)

    csv_content = []
    for row in reader:
        csv_content.append("\t".join(row))
    text_content = "\n".join(csv_content)

    return text_content
