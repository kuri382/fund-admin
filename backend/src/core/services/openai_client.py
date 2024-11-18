import json
from typing import Generator

import openai
from openpyxl import load_workbook

from src.core.services.firebase_driver import AnalysisResult
from src.settings import settings

openai.api_key = settings.openai_api_key


def generate_dd_answer_test(client: openai.ChatCompletion, background: str, dd_prompt):
    response = client.chat.completions.create(
        model="gpt-4-0125-preview",
        messages=[
            {
                "role": "system",
                "content": f"こちらの情報をもとに回答しなさい\n\n{background}",
            },
            {"role": "user", "content": dd_prompt},
        ],
        stream=False,
    )
    return response.choices[0].message.content


def stream_generate_dd_answer_test(client: openai.ChatCompletion, background: str, dd_prompt):
    stream = client.chat.completions.create(
        model="gpt-4-0125-preview",
        messages=[
            {
                "role": "system",
                "content": f"こちらの情報をもとに回答しなさい\n\n{background}",
            },
            {"role": "user", "content": dd_prompt},
        ],
        stream=True,
    )

    for chunk in stream:
        if chunk.choices[0].delta.content is not None:
            yield chunk.choices[0].delta.content


def send_xlsx_content_to_openai(file_path: str, client: openai.ChatCompletion) -> list[str]:
    workbook = load_workbook(file_path)
    sheet = workbook.active
    content = []

    for row in sheet.iter_rows(values_only=True):
        content.append("\t".join([str(cell) for cell in row if cell is not None]))
    text_content = "\n".join(content)

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {
                "role": "user",
                "content": f"次のビジネスで用いられる表の内容に基づいて適切な英語ファイル名のみを回答せよ。拡張子は不要。:\n\n{text_content}",
            }
        ],
        stream=False,
    )
    title = response.choices[0].message.content

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {
                "role": "user",
                "content": f"次のビジネスで用いられる表の概要を1行で短く簡潔説明してください:\n\n{text_content}",
            }
        ],
        stream=False,
    )
    sentence = response.choices[0].message.content

    return title, sentence


def generate_summary(content: str, client: openai.ChatCompletion) -> Generator[str, None, None]:
    prompt = f"以下の内容を基にデューデリジェンス向けのエグゼクティブサマリーを作成してください。日本語でお願いします：\n\n{content}"
    stream = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        stream=True,
    )

    for chunk in stream:
        if chunk.choices[0].delta.content is not None:
            yield chunk.choices[0].delta.content


def generate_market_status(
    content: str,
    client: openai.ChatCompletion,
) -> Generator[str, None, None]:
    prompt = f"以下の内容を基に、市場の状況に関する分析を行ってください。# 市場に対する社内の動向、# 市場に対する社外の動向という二つのセクションから続く形でお願いします。日本語でお願いします。：\n\n{content}"
    stream = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        stream=True,
    )

    for chunk in stream:
        if chunk.choices[0].delta.content is not None:
            yield chunk.choices[0].delta.content


def generate_financial_status(content: str, client: openai.ChatCompletion) -> Generator[str, None, None]:
    prompt = f"以下の内容を基に、財務状況に関する分析を行ってください。日本語でお願いします。：\n\n{content}"
    stream = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        stream=True,
    )

    for chunk in stream:
        if chunk.choices[0].delta.content is not None:
            yield chunk.choices[0].delta.content


def generate_services_status(content: str, client: openai.ChatCompletion) -> Generator[str, None, None]:
    prompt = f"以下の内容を基に、会社の主なサービスや事業に関するレポートを作成してください。会社概要は不要です。日本語でお願いします。：\n\n{content}"
    stream = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        stream=True,
    )

    for chunk in stream:
        if chunk.choices[0].delta.content is not None:
            yield chunk.choices[0].delta.content


def generate_strong_point(content: str, client: openai.ChatCompletion) -> Generator[str, None, None]:
    prompt = f"以下の内容を基にデューデリジェンス資料を作成します。この会社の強みおよびリスクについて、客観的にかつなるべく洞察に富んだ分析をお願いします。IR資料内部の綺麗な表現だけでなく実際に市場で評価されるものなのか分析してください。日本語でお願いします。：\n\n{content}"
    stream = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        stream=True,
    )

    for chunk in stream:
        if chunk.choices[0].delta.content is not None:
            yield chunk.choices[0].delta.content


# unuse
def generate_table_analysis(content: str, openai_client: openai.ChatCompletion) -> Generator[str, None, None]:
    system_prompt = "次のビジネスに関するテーブルデータを分析してください\
        全ての情報を整理して引き出せるように日本語でまとめてください。回答はJSON形式でしてください。\
        abstract: 参照したデータの概要を日本語でまとめてください。一般的な用語説明は不要です\
        extractable_info: このデータから抽出可能な情報の一覧\
        feature: このデータにおいて特徴的な傾向を詳細にまとめてください\
        category:財務会計/管理会計(売上)/管理会計(コスト)/その他 \
        "

    response = openai_client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {'role': 'system', 'content': system_prompt},
            {"role": "user", "content": content},
        ],
        response_format={
            "type": "json_schema",
            "json_schema": {
                "name": "file_analysis",
                "schema": {
                    "type": "object",
                    "properties": {
                        "abstract": {"type": "string"},
                        "feature": {"type": "string"},
                        "extractable_info": {"type": "string"},
                        "category": {"type": "string"},
                    },
                    "required": ["abstract", "feature", "extractable_info", "category"],
                    "additionalProperties": False,
                },
                "strict": True,
            },
        },
    )
    result_raw_abstracts = response.choices[0].message.content
    return result_raw_abstracts


# unuse
def generate_pdf_analysis(content: str, openai_client: openai.ChatCompletion) -> Generator[str, None, None]:
    system_prompt = "次のIRに関わるPDFデータについて分析し日本語でまとめてください。\
        回答はJSON形式でしてください。\
        abstract: 参照したデータの概要を日本語でまとめてください。一般的な用語説明は不要です\
        extractable_info: この資料から抽出可能な情報の一覧\
        feature: このデータにおいて特徴的な傾向を詳細にまとめてください\
        category:ドキュメントの種類について \
        "

    response = openai_client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {'role': 'system', 'content': system_prompt},
            {"role": "user", "content": content},
        ],
        response_format={
            "type": "json_schema",
            "json_schema": {
                "name": "file_analysis",
                "schema": {
                    "type": "object",
                    "properties": {
                        "abstract": {"type": "string"},
                        "feature": {"type": "string"},
                        "extractable_info": {"type": "string"},
                        "category": {"type": "string"},
                    },
                    "required": ["abstract", "feature", "extractable_info", "category"],
                    "additionalProperties": False,
                },
                "strict": True,
            },
        },
    )
    result_raw_abstracts = response.choices[0].message.content
    return json.loads(result_raw_abstracts)


def extract_document_information(
    content_text: str,
    openai_client: openai.ChatCompletion,
) -> AnalysisResult:
    system_prompt = '次のビジネスに関するデータを分析してください。\
        全ての情報を整理して引き出せるように日本語でまとめてください。回答はJSON形式でしてください。\
        abstract: 参照したデータの概要を日本語でまとめてください。一般的な用語説明は不要です。\
        extractable_info: このデータから抽出可能な情報の一覧。\
        feature: このデータにおいて特徴的な傾向を詳細にまとめてください。\
        year_info: このデータが何年何月のデータを示しているか YYYYMM形式で回答せよ\
        period_type: このデータが四半期データか通期データかなどを記述してください。通期/四半期/半期/月次\
        period_number: このデータが第x四半期のデータか int\
        category:財務会計/管理会計(売上)/管理会計(コスト)/その他\
        category_ir: 財務諸表/有価証券報告書/四半期報告書/決算短信または説明資料/適時開示/株主総会招集通知および議決権行使資料/コーポレート・ガバナンス/その他'

    response = openai_client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {'role': 'system', 'content': system_prompt},
            {"role": "user", "content": content_text},
        ],
        response_format={
            "type": "json_schema",
            "json_schema": {
                "name": "file_analysis",
                "schema": {
                    "type": "object",
                    "properties": {
                        "abstract": {
                            "type": "string",
                            "description": "参照したデータの概要を日本語でまとめてください。一般的な用語説明は不要です。",
                        },
                        "feature": {"type": "string"},
                        "extractable_info": {
                            "type": "array",
                            "items": {"type": "string"},
                        },
                        "year_info": {"type": "string"},
                        "period_type": {"type": "string"},
                        "category": {"type": "string"},
                        "category_ir": {"type": "string"},
                    },
                    "required": [
                        "abstract",
                        "feature",
                        "extractable_info",
                        "category",
                        "year_info",
                        "period_type",
                        "category_ir",
                    ],
                    "additionalProperties": False,
                },
                "strict": True,
            },
        },
    )
    result = response.choices[0].message.content
    data = json.loads(result)

    analysis_result = AnalysisResult(
        abstract=data.get('abstract', ''),
        feature=data.get('feature', ''),
        extractable_info=data.get('extractable_info', []),
        year_info=data.get('year_info', ''),
        period_type=data.get('period_type', ''),
        category=data.get('category', ''),
        category_ir=data.get('category_ir', ''),
    )
    return analysis_result


def generate_page_analysis() -> dict:
    response_format = {
        "type": "json_schema",
        "json_schema": {
            "name": "page_analysis",
            "schema": {
                "type": "object",
                "properties": {
                    "abstract": {
                        "type": "string",
                        "description": "参照したデータの概要を日本語でまとめてください。まず始めに結論を書きその後それを捕捉するように文章を構成すること。一般的な用語説明は不要です。",
                    },
                    "extractable_info": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "次の中から抽出可能な情報があれば、選べ。売上高/売上総利益/営業利益/販管費/経常利益/当期純利益/MRR/ARR/社数別単価推移/チャーン/顧客別売上/粗利等",
                    },
                },
                "required": ["abstract", "extractable_info"],
                "additionalProperties": False,
            },
            "strict": True,
        },
    }
    return response_format
