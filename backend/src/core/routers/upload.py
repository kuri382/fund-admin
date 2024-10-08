import io
import os
import json
import uuid
import shutil
from fastapi import APIRouter, Depends, UploadFile, HTTPException, Request
from fastapi.responses import JSONResponse
import openai
from openpyxl import load_workbook
import pandas as pd
import traceback
from typing import TypedDict


from src.core.services.pdf_processing import extract_text_from_pdf
from src.core.services.openai_client import send_xlsx_content_to_openai, generate_file_analysis
from src.core.services.firebase_client import FirebaseClient, get_firebase_client
from src.core.services import auth_service
from src.dependencies import get_openai_client
from src.settings import settings

router = APIRouter()

@router.post("/upload/pdf")
async def upload_file(
    file: UploadFile,
    client: openai.ChatCompletion = Depends(get_openai_client)
):
    file_extension = os.path.splitext(file.filename)[1].lower()

    match file_extension:

        case ".xlsx":
            file_name = os.path.join(settings.table_storage_path, file.filename)
            title, sentence = send_xlsx_content_to_openai(file_name, client)
            new_file_name = os.path.join(settings.table_storage_path, title)
            with open(new_file_name, "wb") as f:
                shutil.copyfileobj(file.file, f)
            return {"filename": file.filename, "status": f"ファイルの分析を完了しました"}

        case ".pdf":
            pdf_text = extract_text_from_pdf(file)
            file_name = os.path.join(settings.pdf_storage_path, f"{file.filename}.txt")
            with open(file_name, 'w') as f:
                f.write(pdf_text)
            return {"filename": file.filename, "status": "PDFからテキスト抽出保存完了"}

        case _:
            raise HTTPException(status_code=400, detail="サポートされていないファイル形式です")


class AnalysisResult(TypedDict):
    abstract: str
    extractable_info: dict
    category: str


@router.post("/upload")
async def upload_file(
    request: Request,
    file: UploadFile,
    firebase_client: FirebaseClient = Depends(get_firebase_client),
    openai_client: openai.ChatCompletion = Depends(get_openai_client),
):
    authorization = request.headers.get("Authorization")
    if not authorization:
        raise HTTPException(status_code=401, detail="Authorization header missing")

    user_id = auth_service.verify_token(authorization)
    file_extension = os.path.splitext(file.filename)[1].lower()

    try:
        contents = await file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File read error: {str(e)}")

    unique_filename = f"{uuid.uuid4()}_{file.filename}"

    match file_extension:

        case ".xlsx":
            try:
                storage_client = firebase_client.get_storage()
                blob = storage_client.blob(f"{user_id}/{unique_filename}")
                blob.upload_from_string(contents, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

                def generate_information(contents, openai_client):
                    file_stream = io.BytesIO(contents)
                    workbook = load_workbook(file_stream)
                    sheet = workbook.active

                    sheet_content = []
                    for row in sheet.iter_rows(values_only=True):
                        sheet_content.append("\t".join([str(cell) for cell in row if cell is not None]))
                    text_content = "\n".join(sheet_content)
                    analysis_result = generate_file_analysis(text_content, openai_client)
                    return json.loads(analysis_result)

                def save_analysis_result(user_id: str, file_name: str, analysis_result: AnalysisResult):
                    firestore_client = firebase_client.get_firestore()
                    doc_ref = firestore_client.collection('analysis_results').document(f"{user_id}_{file_name}")

                    doc_ref.set({
                        "file_name": file_name,
                        "abstract": analysis_result['abstract'],
                        "feature": analysis_result['feature'],
                        "extractable_info": analysis_result['extractable_info'],
                        "category": analysis_result['category']
                    })

                analysis_result = generate_information(contents, openai_client)
                save_analysis_result(user_id, file.filename, analysis_result)

            except Exception as e:
                print(f"Error: {str(e)}")  # エラーメッセージ
                traceback.print_exc()  # 詳細なトレースバックを表示
                raise HTTPException(status_code=500, detail=f"Error uploading Excel file: {str(e)}")

        case _:
            raise HTTPException(status_code=400, detail="サポートされていないファイル形式です")

    return {"filename": file.filename, "status": f"概要：.xlsxとしてリネームし整形・保存しました"}


@router.get("/upload/files")
async def list_excel_files(
    request: Request,
    firebase_client: FirebaseClient = Depends(get_firebase_client),
):
    authorization = request.headers.get("Authorization")
    if not authorization:
        raise HTTPException(status_code=401, detail="Authorization header missing")

    user_id = auth_service.verify_token(authorization)

    try:
        storage_client = firebase_client.get_storage()
        firestore_client = firebase_client.get_firestore()

        blobs = storage_client.list_blobs(prefix=f"{user_id}/")
        file_data_list = []

        for blob in blobs:
            if blob.name.endswith(".xlsx"):
                full_filename = os.path.basename(blob.name)
                filename = full_filename.split("_", 1)[1]

                excel_bytes = blob.download_as_bytes()
                excel_io = io.BytesIO(excel_bytes)

                df = pd.read_excel(excel_io, engine="openpyxl")
                df = df.replace('^Unnamed.*', '', regex=True)
                df = df.fillna('')
                output = df.head(100) # fix later

                output = output.astype(str)
                json_data = output.to_dict(orient="records")

                doc_ref = firestore_client.collection('analysis_results').document(f"{user_id}_{filename}")
                doc = doc_ref.get()

                if doc.exists:
                    analysis_data = doc.to_dict()
                    abstract = analysis_data.get("abstract", "")
                    extractable_info = analysis_data.get("extractable_info", {})
                    category = analysis_data.get("category", "")
                    feature = analysis_data.get("feature", "")
                else:
                    abstract = ""
                    extractable_info = {}
                    category = ""
                    feature = ""

                file_data_list.append({
                    "file_name": filename,
                    "data": json_data,
                    "feature": feature,
                    "abstract": abstract,
                    "extractable_info": extractable_info,
                    "category": category
                })

        if not file_data_list:
            raise HTTPException(status_code=404, detail="No Excel files with numeric data found for the given user.")

        return JSONResponse(content={"files": file_data_list})

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error retrieving or processing files: {str(e)}")



@router.get("/upload/description")
async def list_excel_files(
    request: Request,
    firebase_client: FirebaseClient = Depends(get_firebase_client),
):
    authorization = request.headers.get("Authorization")
    if not authorization:
        raise HTTPException(status_code=401, detail="Authorization header missing")

    user_id = auth_service.verify_token(authorization)

    try:
        firestore_client = firebase_client.get_firestore()

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error retrieving or processing files: {str(e)}")


@router.get("/uploaded-files")
async def get_uploaded_files():
    # ディレクトリ内のファイル名を取得
    files = os.listdir(settings.pdf_storage_path)
    # 拡張子を削除してファイル名を返す
    file_names = [os.path.splitext(file)[0] for file in files if file.endswith('.txt')]
    return {"files": file_names}


@router.get("/uploaded-table-data")
async def get_uploaded_table_data():
    files = os.listdir(settings.table_storage_path)
    file_names = [os.path.splitext(file)[0] for file in files if file.endswith('.txt')]
    return {"files": file_names}
